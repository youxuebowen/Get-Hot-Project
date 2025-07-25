import os

from django.core.paginator import EmptyPage, Paginator
from django.db import transaction
# Create your views here.
from django.shortcuts import render, redirect
from django.contrib import messages
from trio import sleep
from .forms import ExcelUploadForm
from openpyxl import load_workbook
from io import BytesIO
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
from typing import List, Dict, Any, Optional, Tuple
import time
import base64
import os
import pymysql
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse, HttpResponseServerError
from .models import *
import requests
import json
from datetime import datetime, timedelta, timezone  # 日期格式处理
from bs4 import BeautifulSoup  # 爬取内容处理
import re  # 正则表达式
from django.db.models import Q, Sum  # Mysql多条件查询
import smtplib  # 提供了SMTP客户端会话对象，用于向SMTP服务器发送邮件
from email.mime.text import MIMEText  # 用于创建文本类型的邮件内容
import logging

# 获取当前模块的Logger实例
logger = logging.getLogger(__name__)
# 配置信息
TRENDING_API_URL = "https://api.ossinsight.io/q/trending-repos"

GITHUB_API_URL = "https://api.github.com"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36"
GITHUB_TOKEN = None


def index(request):
    return render(request, 'index.html')


# def upload_excel(request):
#     # 上传Excel文件，包含用户的邮箱和校验码
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES['excel_file']
#
#             try:
#                 # 读取Excel文件
#                 wb = load_workbook(filename=BytesIO(excel_file.read()))
#                 sheet = wb.active
#
#                 # 检查表头
#                 headers = [cell.value for cell in sheet[1]]
#                 expected_headers = ['用户名', '邮箱', '校验码']
#                 if headers != expected_headers:
#                     messages.error(request, f'Excel格式不正确，请确保表头为：{", ".join(expected_headers)}')
#                     return redirect('upload')
#                 # 处理数据
#                 # 初始化列表
#                 usernames = []
#                 emails = []
#                 verification_codes = []
#                 success_count = 0
#                 for row in sheet.iter_rows(min_row=2, values_only=True):
#                     username, email, verification_code = row
#                     usernames.append(username)
#                     emails.append(email)
#                     verification_codes.append(verification_code)
#                     success_count += 1
#                 #  发送邮件   默认用第一行的数据发送邮件
#                 if not send_welcome_email(usernames, emails, verification_codes):
#                     error_message = '发送邮件时出错，请检查邮箱配置。'
#                     return redirect(f'/api/v1/fail/?error={error_message}')
#                 messages.success(request, f'成功导入{success_count}条用户数据！')
#                 # 更新数据库
#                 if not update_database():
#                     messages.warning(request, f'更新数据库时出错，请检查数据库配置。')
#                 else:
#                     messages.success(request, f'成功更新数据库！')
#                 return render(request, 'success.html')
#
#             except Exception as e:
#                 messages.error(request, f'处理Excel文件时出错: {str(e)}')
#                 return redirect(f'/api/v1/fail/?error={error_message}')
#             finally:
#                 # 删除文件
#                 if hasattr(excel_file, 'temporary_file_path'):
#                     # 文件存储在临时文件中
#                     file_path = excel_file.temporary_file_path()
#                     if os.path.exists(file_path):
#                         os.remove(file_path)
#
#     else:
#         form = ExcelUploadForm()
#
#     return render(request, 'index.html', {'form': form})

# 搜索查询
def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                # 读取Excel文件
                wb = load_workbook(filename=BytesIO(excel_file.read()))
                sheet = wb.active

                # 检查表头
                headers = [cell.value for cell in sheet[1]]
                expected_headers = ['用户名', '邮箱', '校验码']
                if headers != expected_headers:
                    messages.error(request, f'Excel格式不正确，请确保表头为：{", ".join(expected_headers)}')
                    return redirect('upload')

                # 处理数据
                # 初始化列表
                usernames = []
                emails = []
                verification_codes = []
                success_count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    username, email, verification_code = row
                    usernames.append(username)
                    emails.append(email)
                    verification_codes.append(verification_code)
                    success_count += 1
                #     默认用第一行的数据发送邮件
                if not send_welcome_email(usernames, emails, verification_codes):
                    error_message = '发送邮件时出错，请检查邮箱配置。'
                    return redirect(f'/api/v1/fail/?error={error_message}')
                messages.success(request, f'成功导入{success_count}条用户数据！')
                if not update_database():
                    messages.warning(request, f'更新数据库时出错，请检查数据库配置。')
                else:
                    messages.success(request, f'成功更新数据库！')
                return render(request, 'success.html')

            except Exception as e:
                error_message = f'处理Excel文件时出错: {str(e)}'
                return redirect(f'/api/v1/fail/?error={error_message}')
            finally:
                # 删除文件
                if hasattr(excel_file, 'temporary_file_path'):
                    # 文件存储在临时文件中
                    file_path = excel_file.temporary_file_path()
                    if os.path.exists(file_path):
                        os.remove(file_path)
                # else:
                # return render(request, 'register.html') #文件存储在内存中，无需手动删除
    else:
        form = ExcelUploadForm()

    return render(request, 'index.html', {'form': form})


def content_table_api(request):
    try:
        # 获取查询参数并设置默认值
        page = max(int(request.GET.get('page', 1)), 1)  # 确保至少为1
        page_size = min(max(int(request.GET.get('page_size', 10)), 1), 100)  # 限制10-100条/页
        search = request.GET.get('search', '').strip()
        tag = request.GET.get('tag', '').strip()
        from_date = request.GET.get('from_date', '').strip()
        to_date = request.GET.get('to_date', '').strip()
        content = request.GET.get('content', '').strip()
        type_value = request.GET.get('type', '').strip()

        # 使用select_related/prefetch_related优化关联查询（如果有外键）
        queryset = HotProjects.objects.all().order_by('-updated_time')

        # 应用搜索过滤
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(tag__icontains=search) |
                Q(description__icontains=search)
            )

        # 应用内容过滤
        if content:
            queryset = queryset.filter(
                Q(content__icontains=content) |
                Q(name__icontains=content)
            )

        # 应用标签过滤（支持多标签，使用OR连接）
        if tag:
            tag_list = [t.strip() for t in tag.split(',') if t.strip()]
            if tag_list:
                tag_query = Q()
                for t in tag_list:
                    tag_query |= Q(tag__icontains=t)
                queryset = queryset.filter(tag_query)

        # 应用类型过滤
        if type_value and type_value.isdigit():
            queryset = queryset.filter(type=int(type_value))

        # 应用时间范围过滤
        try:
            if from_date:
                # 将字符串日期转换为datetime对象
                from_datetime = datetime.datetime.strptime(from_date, '%Y-%m-%d')
                queryset = queryset.filter(created_time__gte=from_datetime)

            if to_date:
                # 将字符串日期转换为datetime对象，并设置为当天结束时间
                to_datetime = datetime.datetime.strptime(to_date, '%Y-%m-%d')
                to_datetime = to_datetime.replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(created_time__lte=to_datetime)
        except ValueError as e:
            # 日期格式错误
            return JsonResponse({'error': f'日期格式错误: {str(e)}'}, status=400)

        # 性能关键：避免在分页前执行count()
        # 使用Paginator的懒加载特性
        paginator = Paginator(queryset, page_size)

        try:
            current_page = paginator.page(page)
        except EmptyPage:
            # 超出范围时返回最后一页
            current_page = paginator.page(paginator.num_pages)

        # 构建响应数据 - 使用values()优化序列化
        results = list(current_page.object_list.values(
            'id', 'name', 'description', 'url',
            'tag', 'type', 'if_sent', 'updated_time'
        ))

        # 转换字段格式
        for item in results:
            item['status'] = '已经推送' if item['if_sent'] else '未被推送'
            item['updated_time'] = item['updated_time'].isoformat() if item['updated_time'] else None
            del item['if_sent']  # 移除原始字段

        data = {
            'count': paginator.count,
            'next': current_page.next_page_number() if current_page.has_next() else None,
            'previous': current_page.previous_page_number() if current_page.has_previous() else None,
            'total_pages': paginator.num_pages,
            'current_page': current_page.number,
            'results': results
        }

        return JsonResponse(data)

    except Exception as e:
        # 使用正式日志系统替代print
        # logger.error(f"API错误: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'服务器内部错误: {str(e)}'}, status=500)


def update_chosen_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            if_chosen = data.get('if_chosen', 1)  # 默认设为1
            if not ids:
                return JsonResponse({'success': False, 'error': '没有提供ID列表'})

            # 更新选中条目的if_chosen字段
            updated_count = HotProjects.objects.filter(id__in=ids).update(if_chosen=if_chosen)

            return JsonResponse({
                'success': True,
                'updated_count': updated_count
            })
        except Exception as e:
            print(f"更新if_chosen错误: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'error': '不支持的请求方法'}, status=405)


def get_chosen_content_api(request):
    try:
        # 获取查询参数
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))

        # 基础查询集 - 只查询is_chosen=1的条目
        queryset = HotProjects.objects.filter(if_chosen=1).order_by('-updated_time')

        # 分页处理
        paginator = Paginator(queryset, page_size)
        current_page = paginator.get_page(page)

        # 构建响应数据
        data = {
            'count': paginator.count,
            'next': current_page.next_page_number() if current_page.has_next() else None,
            'previous': current_page.previous_page_number() if current_page.has_previous() else None,
            'results': [{
                'id': item.id,
                'name': item.name,
                'description': item.description,
                'url': item.url,
                'tag': item.tag,
                'type': item.type,
                'status': '已经推送' if item.if_sent else '未被推送',
                'updated_time': item.updated_time.isoformat() if item.updated_time else None
            } for item in current_page.object_list]
        }

        return JsonResponse(data)
    except Exception as e:
        print(f"获取已选内容错误: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def fail(request):
    return render(request, 'fail.html')


def success(request):
    return render(request, 'success.html')


# def query_hot_projects(request, self=None):
#     if request.method == 'GET':
#         content = request.GET.get('content')
#         tag = request.GET.get('tag')
#         article_type = request.GET.get('article_type')
#         if article_type is not None:
#             article_type = int(article_type)
#         from_date = request.GET.get('from_date')
#         to_date = request.GET.get('to_date')
#         sort_order = request.GET.get('sort_order', 'desc')
#         page = int(request.GET.get('page', 1))
#         page_size = int(request.GET.get('page_size', 10))
#         query_tool = TechArticleQueryTool(
#             host='localhost',
#             user='root',
#             password='123456',
#             database='sky_take_out'
#         )
#         data = query_tool.query_articles(content=content, tag=tag, article_type=article_type,
#                        from_date=from_date, to_date=to_date, sort_order=sort_order,
#                        page=page, page_size=page_size)
#         results = []
#         for item in data:
#             result = {
#                 "id": item["id"],
#                 "name": item["name"],
#                 "description": item["description"],
#                 "url": item["url"],
#                 "tag": item["tag"],
#                 "type": item["type"],
#                 "status": '已经推送' if item["if_sent"] == 1 else '未被推送',
#                 "updated_time": item["updated_time"]
#             }
#             results.append(result)
#         return JsonResponse({"results": results}, safe=False, json_dumps_params={'ensure_ascii': False})
#         # result = {
#         #     "id": data["id"],
#         #     "name": data["name"],
#         #     "description": data["description"],
#         #     "url":data["url"],
#         #     "tag":data["tag"],
#         #     "type":data["type"],
#         #     "status": '已经推送' if data["if_sent"]==1 else '未被推送',
#         #     "updated_time":data["updated_time"]
#         # }
#         # result = json.dumps(result)
#         # return JsonResponse(result)

def get_article_url(request):
    # 获取热门文章的链接，榜单三天刷一次，每三天运行一次就行了
    articles = get_artical_link()
    if insert_articles(articles):
        return JsonResponse({"message": "文章数据已成功获取并插入数据库"}, json_dumps_params={'ensure_ascii': False})
    else:
        return JsonResponse({"message": "插入失败"}, json_dumps_params={'ensure_ascii': False})


def get_article_descriptions(request):
    # 更新description和tag
    sql_links = read_articles_sql()  # 获取description为空的记录,且type=2是掘金类型
    articles_tag, articles_description = get_articles_description_tag(sql_links)  # 获取这些记录的 description
    update_articles_descriptions(sql_links, articles_tag, articles_description)  # 更新对应记录
    return JsonResponse({"message": f"成功更新 {len(articles_description)} 条文章描述及标签"},
                        json_dumps_params={'ensure_ascii': False})


# 获取github的url链接并存储到数据库
def github_url(request):
    period = request.GET.get('period', 'past_week')
    data_count = request.GET.get('data_count', 5)  # 每次爬50条
    language = request.GET.get('language', 'Python')

    # 限制爬取的条数在0-100
    if not str(data_count).isdigit() or int(data_count) <= 0:
        return JsonResponse({"code": 400, "message": "data_count must be positive"})
    data_count = min(int(data_count), 100)
    #
    try:

        # params = {'period': period, 'data_count': data_count}
        # response = requests.get(TRENDING_API_URL, params = params, headers={'user-agent': USER_AGENT})
        max_retries = 3
        response = requests.get(
            TRENDING_API_URL,
            headers={'User-Agent': USER_AGENT},
            params={'period': period, 'data_count': data_count, 'language': language}  # 显式传递参数
        )
        for attempt in range(max_retries):
            if response.status_code == 200:
                break
            time.sleep(1)  # 延迟1秒后重试
        else:
            return JsonResponse({"code": 503, "message": "API暂时不可用，请稍后重试"})
        response.raise_for_status()
        data = response.json()
        repos = []

        for item in data.get('data', [])[:data_count]:
            repo_name = item['repo_name']
            repo_url = f"https://github.com/{repo_name}"

            """
            数据库查询部分，查询该项目是否已经存在，若不存在（重复数量为0），则进行插入
            """
            """start"""
            urlIfExist = HotProjects.objects.filter(name=item['repo_name']).count()
            if urlIfExist == 0:
                # 存储/更新数据库
                HotProjects.objects.update_or_create(
                    name=repo_name,
                    defaults={
                        'name': repo_name,
                        'url': repo_url,
                        'type': 1,
                        'if_sent': 0,
                        'if_chosen': 0,
                        'created_time': time.time()
                    }
                )
                """end"""

                repos.append({'name': repo_name, 'url': repo_url})
        # sync_readme_to_db()
        return JsonResponse({"code": 200, "data": repos})
    except Exception as e:
        return JsonResponse({"code": 500, "message": str(e)})


# 获取数据库中content为空的url，调用fetch_readme_content去爬取，根据返回内容存入数据库
def save_github_readme(request):
    # 改了name
    github_repo_name = HotProjects.objects.filter(content="").values('name')
    return_data = []
    for repo_name in github_repo_name:
        # repo_name = request.GET.get('repo_name')
        try:
            content = fetch_readme_content(repo_name)
            tag = agent(content)

            # 更新数据库
            """爬取数据跟新，插入readme的content、tag、description字段"""
            """start"""
            HotProjects.objects.filter(name=repo_name['name']).update(content=content, description=tag[1], tag=tag[0],
                                                                      updated_time=datetime.now())
            """end"""

            return_data.append([content, tag])
            # return JsonResponse({"code": 200, "data": tag})
        except Exception as e:
            return JsonResponse({"code": 500, "message": str(e), "data": return_data})

    return return_data


"""
获取漏洞信息
"""


def cve_email_send(request):
    now = datetime.now().date()
    cve_id_list = CveSpider.objects.filter(created_time__gte=now).values("cve_id")
    description_cn_list = CveSpider.objects.filter(created_time__gte=now).values("description_cn")
    url_list = CveSpider.objects.filter(created_time__gte=now).values("url")
    # published_time_list = CveSpider.objects.filter(created_time__gte=now).values("published_time")

    auth = request.GET['auth']
    sender = request.GET['sender']
    mailserver = "smtp.js.chinamobile.com"

    send_list = ["jinyiran@js.chinamobile.com", "352903859@qq.com"]

    text_content = "今日CVE漏洞披露信息：\n"
    for i in range(len(cve_id_list)):
        text_content += "· " + cve_id_list[i]["cve_id"] + ":\n"
        text_content += "漏洞简介: " + description_cn_list[i]["description_cn"] + "\n"
        # text_content += "披露时间: " + published_time_list[i]["published_time"] + "\n"
        text_content += "原始链接: " + url_list[i]["url"] + "\n"
        text_content += "\n\n"
    """
    url = "https://zhenze-huhehaote.cmecloud.cn/v1/chat/completions"
    api_key = "aJophUtdoVendosyMJIWcuaDTtcaBmpBrMh7Y8EvDRg"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    data = {
            "model": "Qwen2.5-72B-Instruct-64K",
            "messages": [
                {
                    "role": "user",
                    "content": "以下内容为我搜集的当日互联网上公开的CVE安全漏洞信息，请将此内容整理为一封便于阅读的简报，返回的内容前后不要添加问候语、祝好、署名等任何与我给你的内容无关的东西: " + text_content
                },
                {
                    "role": "system",
                    "content": ""
                }
            ],
            "max_tokens": 16384,
            "stream": False,
            "temperature": 0.7,
            "top_p": 0.9
        }
    response = requests.post(url, headers=headers, json=data)
    if response.status_code == 200:
        response.encoding = "utf-8"
        email_content = json.loads(response.text)['choices'][0]['message']['content']
        #print(description)
    else:
        print(f"请求失败，状态码: {response.status_code}")
        print(response.text)
    """

    # 发送邮件
    subject = '【自动发送】今日CVE漏洞披露信息'
    message = MIMEText(text_content, 'plain', 'utf-8')
    message['From'] = sender  # 设置发件人信息
    message['To'] = ','.join(send_list)  # 设置收件人信息，多个收件人用逗号分隔
    message['Subject'] = subject  # 设置邮件主题
    smtp = smtplib.SMTP_SSL(mailserver, 465)
    smtp.login(sender, auth)
    smtp.sendmail(sender, send_list, message.as_string())
    smtp.quit()

    return JsonResponse({"code": 200, "message": "Successful"})


def cve_info_list(request):
    if request.GET.get('cveId') == None or request.GET.get('pageFrom') == None or request.GET.get(
            'pageSize') == None or request.GET.get('dateFrom') == None or request.GET.get('dateTo') == None:
        return JsonResponse({"code": 500, "message": "Some parameter is missed"})
    elif request.GET.get('cveId') == "all" and request.GET.get('dateFrom') != "all" and request.GET.get(
            'dateTo') != "all":
        dateFrom = datetime.strptime(request.GET['dateFrom'], '%Y%m%d').strftime('%Y-%m-%d %H:%M:%S')
        dateTo = datetime.strptime(request.GET['dateTo'], '%Y%m%d').strftime('%Y-%m-%d %H:%M:%S')
        pageSize = int(request.GET['pageSize'])
        pageFrom = int(request.GET['pageFrom'])
        offset = (pageFrom - 1) * pageSize
        result_list = CveSpider.objects.filter(created_time__gte=dateFrom, created_time__lte=dateTo).order_by(
            '-created_time').values()[offset:offset + pageSize]
        result_num = CveSpider.objects.filter(created_time__gte=dateFrom, created_time__lte=dateTo).order_by(
            '-created_time').count()
        return JsonResponse({"code": 200, "message": "Successful",
                             "data": {"result_num": result_num, "detail": list(result_list.values())}})
    elif request.GET.get('cveId') != "all" and request.GET.get('dateFrom') != "all" and request.GET.get(
            'dateTo') != "all":
        dateFrom = datetime.strptime(request.GET['dateFrom'], '%Y%m%d').strftime('%Y-%m-%d %H:%M:%S')
        dateTo = datetime.strptime(request.GET['dateTo'], '%Y%m%d').strftime('%Y-%m-%d %H:%M:%S')
        pageSize = int(request.GET['pageSize'])
        pageFrom = int(request.GET['pageFrom'])
        offset = (pageFrom - 1) * pageSize
        result_list = CveSpider.objects.filter(
            Q(cve_id=request.GET['cveId']) & Q(created_time__gte=dateFrom, created_time__lte=dateTo)).order_by(
            '-created_time').values()[offset:offset + pageSize]
        result_num = CveSpider.objects.filter(
            Q(cve_id=request.GET['cveId']) & Q(created_time__gte=dateFrom, created_time__lte=dateTo)).order_by(
            '-created_time').count()
        return JsonResponse({"code": 200, "message": "Successful",
                             "data": {"result_num": result_num, "detail": list(result_list.values())}})
    elif request.GET.get('cveId') != "all" and request.GET.get('dateFrom') == "all" and request.GET.get(
            'dateTo') == "all":
        pageSize = int(request.GET['pageSize'])
        pageFrom = int(request.GET['pageFrom'])
        offset = (pageFrom - 1) * pageSize
        result_list = CveSpider.objects.filter(cve_id=request.GET['cveId']).order_by('-created_time').values()[
                      offset:offset + pageSize]
        result_num = CveSpider.objects.filter(cve_id=request.GET['cveId']).order_by('-created_time').count()
        return JsonResponse({"code": 200, "message": "Successful",
                             "data": {"result_num": result_num, "detail": list(result_list.values())}})
    elif request.GET.get('cveId') == "all" and request.GET.get('dateFrom') == "all" and request.GET.get(
            'dateTo') == "all":
        pageSize = int(request.GET['pageSize'])
        pageFrom = int(request.GET['pageFrom'])
        offset = (pageFrom - 1) * pageSize
        result_list = CveSpider.objects.order_by('-created_time').values()[offset:offset + pageSize]
        result_num = CveSpider.objects.order_by('-created_time').count()
        return JsonResponse({"code": 200, "message": "Successful",
                             "data": {"result_num": result_num, "detail": list(result_list.values())}})
    else:
        return JsonResponse({"code": 500, "message": "Internal Error"})


def cve_num_list(request):
    now = datetime.now().date()
    yesterday = now - timedelta(days=1)
    now = datetime.now().date().strftime('%Y-%m-%d %H:%M:%S')
    yesterday = yesterday.strftime('%Y-%m-%d %H:%M:%S')

    cve_total_num = CveSpider.objects.count()
    cve_yesterday_num = CveSpider.objects.filter(created_time__gte=yesterday, created_time__lte=now).count()
    return JsonResponse({"code": 200, "message": "Successful",
                         "data": {"cve_total_num": cve_total_num, "cve_yesterday_num": cve_yesterday_num}})


# 定义发送邮件的函数
def send_welcome_email(usernames, Emails, verification_codes):
    """
    发送欢迎邮件给新用户

    :param username: 用户姓名
    :param email: 用户邮箱
    :param verification_code: 用户校验码
    :return: 发送成功返回 True，失败返回 Falseauth_user
    """
    mailserver = 'smtp.qq.com'
    subject = '【WideSeek】每周技术热点推送！'
    userName_SendMail = Emails[0]
    userName_AuthCode = verification_codes[0]
    projects = get_hot_projects()
    for username, Email, verification_code in zip(usernames[1:], Emails[1:], verification_codes[1:]):
        message = ""
        for project in projects:
            message += f'热点标题：{project["name"]}\n\n内容速览:\n {project["description"]}\n链接地址: {project["url"]}\n\n\n'
            # get_tage_projects(content=project["content"])
        message = f"尊敬的 {username}，以下是本周技术热点推送：\n\n{message}"
        email = MIMEText(message, "plain", "utf-8")
        email["Subject"] = subject  # 定义邮件主题
        email["From"] = userName_SendMail  # 发件人
        email["To"] = ",".join(
            Emails[1:]
        )  # 收件人（可以添加多个，若只有一个收件人，可直接写邮箱号）
        try:
            smtp = smtplib.SMTP_SSL(mailserver, 465)
            smtp.login(userName_SendMail, userName_AuthCode)
            smtp.sendmail(userName_SendMail, Email, email.as_string())
            smtp.quit()
        except Exception as e:
            print(f'给 {Emails[1:]} 发送邮件时出错: {str(e)}')
            return False
    return True


# 更新数据库，清空待发送列表，状态改为已发送
def update_database():
    HotProjects.objects.filter(if_chosen=1).update(if_chosen=0, if_sent=1)
    return True


# 获取数据库中的待发送项目,if_chosen=1
def get_hot_projects():
    # 使用 filter 方法筛选 if_send 为 1 的记录，并只获取 name 和 字段description
    projects = HotProjects.objects.filter(if_chosen=1).values('name', 'description', 'url')
    return projects


# 获取文章列表信息，返回列表/字典/
def get_artical_link():
    url_artical_link = "https://api.juejin.cn/content_api/v1/content/article_rank?category_id=1&type=hot"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Referer": "https://juejin.cn/",
    }

    try:
        response = requests.get(url_artical_link, headers=headers, timeout=10)
        response.raise_for_status()

        data = response.json()

        # 检查API返回状态
        if data.get("err_no") != 0 or "data" not in data:
            print(f"API返回错误: {data.get('err_msg', '未知错误')}")
            return []

        # articles = data["data"][:50]
        # 每次爬取二十个url
        articles = data["data"][:20]
        results = []
        for article in articles:
            # 确保content字段存在
            if "content" not in article:
                continue

            content = article["content"]
            title = content.get("title", "无标题")
            article_id = content.get("content_id", "")

            # 构建完整URL
            article_url = f"https://juejin.cn/post/{article_id}" if article_id else ""

            results.append({
                "url": article_url,
                "name": title,
                "content": "",
                "description": "",
                "tag": " ",
                "type": 2,
                "if_sent": 0,
            })

        return results
    except Exception as e:
        print(f"爬取失败: {str(e)}")
        return []
# 将文章url数据存入库中
def insert_articles(articles):
    new_count = 0
    existing_urls = set(HotProjects.objects.values_list('url', flat=True))
    for article in articles:
        if article["url"] in existing_urls:
            continue
        try:
            HotProjects.objects.create(
                url=article["url"],
                name=article["name"],
                type=article["type"],
                if_sent=article["if_sent"],
                if_chosen=0,
            )
            new_count += 1
            existing_urls.add(article["url"])
        except Exception as e:
            print(f"插入数据失败: {e}")
            return False
    print(f"成功新增 {new_count} 条数据")
    return True

# 获取description为空的记录,且type=2是掘金类型
def read_articles_sql():
    results = []
    # 使用 ORM 查询 description 字段为空的记录，限制为三个防止爬取不回来
    limit_count = 1
    articles = HotProjects.objects.filter(description='').values('id', 'url', 'name', 'type', 'if_sent')[:limit_count]
    # 转换为字典格式
    for row in articles:
        results.append({
            'id': row['id'],
            'url': row['url'],
            'name': row['name'],
            'type': row['type'],
            'if_sent': row['if_sent'],
        })

    return results

# 向智能体发送请求，获取description,tag返回的是列表/字典
# 实际github也可以调用
def get_articles_description_tag(informations):
    articles_description = []
    articles_tag = []
    a = 1
    headers_list = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Referer": "https://juejin.cn/",
    }
    url_3 = 'https://yuanqi.tencent.com/openapi/v1/agent/chat/completions'
    token_1 = "P0DLg1Kduqe5jIDQvSyTsqc9vhmhp0El"
    assistant_id_1 = "ZvxqWUOM05De"
    # 获取 tag
    token_2 = "XAoYVGiSgGADDDmzKd7IQC1uI6bOWCnB"
    assistant_id_2 = "GyUeYNOLbmaq"
    # 获取摘要
    token_3 = "PrWV3i71LuSEJ9cSkK9MuhNXe3b8Za0Z"
    assistant_id_3 = "mqm8WNPSOBi7"
    token_4 = "sfl5dL6WZGH2X4IK5w9NBNZOQ2Se0tDH"
    assistant_id_4 = "eA3znc5W7lRs"
    token = token_4
    assistant_id = assistant_id_4

    for info in informations:
        link = info['url']
        try:
            try:
                # 记录不同级别的日志
                logger.debug("这是一条DEBUG日志（调试信息）")
                # 获取网页内容
                # 使用 Selenium 获取网页内容
                # page_source = get_page_content(link)
                # if page_source is None:
                #     continue
                # soup = BeautifulSoup(page_source, features='html.parser')
                # 获取网页内容
                res = requests.get(link, headers=headers_list, timeout=10)
                res.raise_for_status()  # 检查请求是否成功
                soup = BeautifulSoup(res.text, features='html.parser')
            except Exception as e:
                logger.error("发生错误: %s", str(e), exc_info=True)  # 记录异常堆栈
                return HttpResponseServerError("未爬到具体内容")

            # 调用 API
            headers = {
                'X-Source': 'openapi',
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {token}',
            }
            data_1 = {
                "assistant_id": f"{assistant_id}",
                "user_id": "username",
                "stream": False,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": soup.text[:100000],  # 限制文本长度，避免超出 API 限制
                            }
                        ]
                    }
                ]
            }

            # 发送请求到智能体 API
            api_res = requests.post(url_3, headers=headers, json=data_1, timeout=30)
            api_res.raise_for_status()  # 检查 API 请求是否成功

            # 解析 JSON 响应
            try:
                data_artical = api_res.json()
                content_description = data_artical['choices'][0]["message"]["content"]
                description_tag = content_description.rsplit('#', 1)
                articles_description.append({"description": description_tag[1]})
                articles_tag.append({"tag": description_tag[0]})
                print(f"成功获取链接 {link} 的描述")

            except (KeyError, IndexError) as e:
                print(f"解析 API 响应失败: {e}")
                print(f"API 返回内容: {api_res.text[:500]}")  # 打印部分响应内容用于调试
            except requests.exceptions.JSONDecodeError:
                print(f"API 返回非 JSON 格式内容: {api_res.text[:500]}")

        except requests.exceptions.RequestException as e:
            print(f"处理链接 {link} 时出错: {e}")

        # 限制处理数量（测试用）
    return articles_tag, articles_description

def get_page_content(url):
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # 无头模式，适用于服务器环境
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')

    # 请根据服务器环境修改 ChromeDriver 的路径
    service = Service('/path/to/chromedriver')
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        driver.get(url)
        # 等待页面加载完成
        time.sleep(5)
        page_source = driver.page_source
        return page_source
    except Exception as e:
        print(f"获取页面内容失败: {e}")
        return None
    finally:
        driver.quit()
# 数据库更新description
def update_articles_descriptions(sql_links, articles_tag, articles_descriptions):
    for sql_link, articles_tag, articles_description in zip(sql_links, articles_tag, articles_descriptions):
        article_id = sql_link['id']
        description = articles_description["description"]
        tag = articles_tag["tag"]
        try:
            article = HotProjects.objects.get(id=article_id)
            article.description = description
            article.tag = tag
            article.save()
            print(f"成功更新 ID 为 {article_id} 的记录的 description 和 tag")
        except HotProjects.DoesNotExist:
            print(f"ID 为 {article_id} 的记录不存在")
    return True

# 获取tag为none
def read_articles_sql_tag():
    results = []
    # 使用 ORM 查询 tag 字段为空的记录
    articles = HotProjects.objects.filter(tag='').values('id', 'url', 'name', 'type', 'if_sent')

    # 转换为字典格式
    for row in articles:
        results.append({
            'id': row['id'],
            'url': row['url'],
            'name': row['name'],
            'type': row['type'],
            'if_sent': row['if_sent']
        })

    return results

# 调用智能体获取tag
def get_articles_tag(informations):
    articles_tag = []
    headers_list = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Referer": "https://juejin.cn/",
    }
    url_3 = 'https://yuanqi.tencent.com/openapi/v1/agent/chat/completions'
    token_1 = "P0DLg1Kduqe5jIDQvSyTsqc9vhmhp0El"
    assistant_id_1 = "ZvxqWUOM05De"
    # 获取tag
    token_2 = "XAoYVGiSgGADDDmzKd7IQC1uI6bOWCnB"
    assistant_id_2 = "GyUeYNOLbmaq"
    # 获取摘要
    token_3 = "PrWV3i71LuSEJ9cSkK9MuhNXe3b8Za0Z"
    assistant_id_3 = "mqm8WNPSOBi7"
    token = token_2
    assistant_id = assistant_id_2

    for info in informations:
        sleep(20)
        link = info['url']
        try:
            # 获取网页内容
            res = requests.get(link, headers=headers_list, timeout=10)
            res.raise_for_status()  # 检查请求是否成功
            soup = BeautifulSoup(res.text, features='html.parser')

            # 调用API
            headers = {
                'X-Source': 'openapi',
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {token}',
            }
            data_1 = {
                "assistant_id": f"{assistant_id}",
                "user_id": "username",
                "stream": False,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": soup.text[:100000],  # 限制文本长度，避免超出API限制
                            }
                        ]
                    }
                ]
            }

            # 发送请求到智能体API
            api_res = requests.post(url_3, headers=headers, json=data_1, timeout=30)
            api_res.raise_for_status()  # 检查API请求是否成功

            # 解析JSON响应
            try:
                data_artical = api_res.json()
                content_tag = data_artical['choices'][0]["message"]["content"]
                articles_tag.append({"tag": content_tag})
                print(f"成功获取链接 {link} 的tag")
            except (KeyError, IndexError) as e:
                print(f"解析API响应失败: {e}")
                print(f"API返回内容: {api_res.text[:500]}")  # 打印部分响应内容用于调试
            except requests.exceptions.JSONDecodeError:
                print(f"API返回非JSON格式内容: {api_res.text[:500]}")

        except requests.exceptions.RequestException as e:
            print(f"处理链接 {link} 时出错: {e}")

        # 限制处理数量（测试用）
        # if a > 4:
        #     break

    return articles_tag  # 将description存入库中，更新数据库

# 数据库更新tag
def update_articles_tag(sql_links, articles):
    try:
        # 使用 Django 的事务上下文管理器确保数据一致性
        with transaction.atomic():
            for sql_link, article in zip(sql_links, articles):
                article_id = sql_link['id']
                tag = article["tag"]
                # 使用 ORM 更新指定 ID 记录的 tag 字段
                HotProjects.objects.filter(id=article_id).update(tag=tag)
            print(f"成功更新 {len(articles)} 条记录的 tag")
    except Exception as err:
        print(f"更新数据失败: {err}")
"""
暂时没用
负责搜索 GitHub 仓库、提取 README 内容和处理仓库信息。
包含以下主要方法：
search_repositories：根据关键词搜索 GitHub 仓库。
get_readme_content：获取指定仓库的 README 内容，有重试机制处理速率限制。
extract_repo_info：从 GitHub URL 提取仓库信息。
find_readme_url：在仓库页面查找 README.md 链接。
process_repositories：处理仓库搜索和 README 提取流程，可选择将结果保存到数据库。
"""


# 自定义JsonResponse
# class JsonResponse(HttpResponse):
#     def __init__(self, data, **kwargs):
#         content = json.dumps(data, ensure_ascii=False)
#         kwargs.setdefault('content_type', 'application/json')
#         super().__init__(content, **kwargs)
# 调用智能体，获取标签和描述
def agent(text):
    url = 'https://yuanqi.tencent.com/openapi/v1/agent/chat/completions'

    headers = {
        'X-Source': 'openapi',
        'Content-Type': 'application/json',
        'Authorization': 'Bearer sfl5dL6WZGH2X4IK5w9NBNZOQ2Se0tDH'
    }

    # 构建工作流请求体
    data = {
        "assistant_id": "eA3znc5W7lRs",
        "user_id": "username",
        "stream": False,
        "messages": [{
            "role": "user",
            "content": [{
                "type": "text",
                "text": text
            }]
        }]
    }

    response = requests.post(url, headers=headers, json=data)

    response_data = response.json()
    # print(response.text)
    content_description = response_data['choices'][0]["message"]["content"]
    description_tag = content_description.rsplit('#', 1)
    # print(description_tag)
    return description_tag

# 爬取README
def fetch_readme_content(repo_name):
    # 因为是字典，不能直接用repo_name，需要用repo_name['name']
    url = f"{GITHUB_API_URL}/repos/{repo_name['name']}/readme"
    headers = {'User-Agent': USER_AGENT}
    if GITHUB_TOKEN:
        headers['Authorization'] = f'token {GITHUB_TOKEN}'
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        if 'content' in data:
            content = base64.b64decode(data['content']).decode('utf-8')
            # tag = agent(content)
            return content
        return "README内容不可用"
    except Exception as e:
        return f"获取失败: {str(e)}"


# 盒子/图表函数
def project_tags_api(request):
    try:
        # 只查询type为1的文章数据
        articles = HotProjects.objects.filter(type=1)

        # 标签统计字典
        tag_counts = {}

        # 定义有效的标签列表 - 确保包含所有需要统计的标签
        valid_tags = {
            "后端", "Java", "Go", "Kotlin", "Python", "Node.js", "Swift", "前端",
            "JavaScript", "TypeScript", "Vue.js", "React.js", "HTML", "CSS", "three.js",
            "Flutter", "Spring Boot", "Spring", "Android", "Jetpack", "Trae", "Cursor",
            "MySQL", "数据库", "架构", "Linux", "HarmonyOS", "云原生", "云计算", "算法",
            "Debug", "逆向", "源码", "开源", "编程语言", "AI编程", "爬虫"
        }

        # 遍历所有文章
        for article in articles:
            # 确保tag字段存在且不为空
            if article.tag:
                # 分割标签并去除空格
                # 处理可能的多种分隔符（逗号、空格等）
                tags = [tag.strip() for tag in article.tag.replace('，', ',').split(',') if tag.strip()]

                # 统计每个有效标签
                for tag in tags:
                    # 转换为小写进行匹配，确保大小写不敏感
                    lower_tag = tag.lower()
                    # 查找有效的标签（不区分大小写）
                    matched_tag = next((t for t in valid_tags if t.lower() == lower_tag), None)
                    if matched_tag:
                        tag_counts[matched_tag] = tag_counts.get(matched_tag, 0) + 1
                    else:
                        # 对于不在valid_tags中的标签，也进行统计，方便调试
                        print(f"未匹配的标签: {tag}")

        # 转换为列表并按数量排序
        tag_list = [{'name': tag, 'value': count} for tag, count in tag_counts.items()]
        tag_list.sort(key=lambda x: x['value'], reverse=True)

        # 返回结果
        return JsonResponse({'tags': tag_list})
    except Exception as e:
        print(f"project标签统计API错误: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def article_tags_api(request):
    try:
        # 只查询type为2的文章数据
        articles = HotProjects.objects.filter(type=2)

        # 标签统计字典
        tag_counts = {}

        # 定义有效的标签列表 - 确保包含所有需要统计的标签
        valid_tags = {
            "后端", "Java", "Go", "Kotlin", "Python", "Node.js", "Swift", "前端",
            "JavaScript", "TypeScript", "Vue.js", "React.js", "HTML", "CSS", "three.js",
            "Flutter", "Spring Boot", "Spring", "Android", "Jetpack", "Trae", "Cursor",
            "MySQL", "数据库", "架构", "Linux", "HarmonyOS", "云原生", "云计算", "算法",
            "Debug", "逆向", "源码", "开源", "编程语言", "AI编程", "爬虫"
        }

        # 遍历所有文章
        for article in articles:
            # 确保tag字段存在且不为空
            if article.tag:
                # 分割标签并去除空格
                # 处理可能的多种分隔符（逗号、空格等）
                tags = [tag.strip() for tag in article.tag.replace('，', ',').split(',') if tag.strip()]

                # 统计每个有效标签
                for tag in tags:
                    # 转换为小写进行匹配，确保大小写不敏感
                    lower_tag = tag.lower()
                    # 查找有效的标签（不区分大小写）
                    matched_tag = next((t for t in valid_tags if t.lower() == lower_tag), None)
                    if matched_tag:
                        tag_counts[matched_tag] = tag_counts.get(matched_tag, 0) + 1
                    else:
                        # 对于不在valid_tags中的标签，也进行统计，方便调试
                        print(f"未匹配的标签: {tag}")

        # 转换为列表并按数量排序
        tag_list = [{'name': tag, 'value': count} for tag, count in tag_counts.items()]
        tag_list.sort(key=lambda x: x['value'], reverse=True)

        # 返回结果
        return JsonResponse({'tags': tag_list})
    except Exception as e:
        print(f"标签统计API错误: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def stats_data_api(request):
    try:
        # 获取今日和昨日的日期
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)
        first_day_of_month = today.replace(day=1)

        # 从HotProjects获取数据
        # hotArticlesTotal: type为2的文章个数
        hot_articles_total = HotProjects.objects.filter(type=2).count()

        # hotProjectsTotal: type为1的文章条目总个数
        hot_projects_total = HotProjects.objects.filter(type=1).count()

        # pushedTotal: if_sent字段为1的所有文章个数
        pushed_total = HotProjects.objects.filter(if_sent=1).count()

        # toPushTotal: if_sent字段为0且if_chosen字段为1的所有文章个数
        to_push_total = HotProjects.objects.filter(if_sent=0, if_chosen=1).count()

        # 从CveSpider获取数据
        # vulnerabilitiesTotal: 数据库总条目个数
        vulnerabilities_total = CveSpider.objects.count()

        # yesterdayVulns: 今日新增的条目个数
        today_cve = CveSpider.objects.filter(created_time__date=today).count()

        # 计算hotArticlesChange
        today_articles = HotProjects.objects.filter(type=2, created_time__date=today).count()
        yesterday_articles = HotProjects.objects.filter(type=2, created_time__date=yesterday).count()

        if yesterday_articles == 0:
            hot_articles_change = "较昨日 ↑ 100%" if today_articles > 0 else "较昨日不变"
        else:
            rate = ((today_articles - yesterday_articles) / yesterday_articles) * 100
            if rate > 0:
                hot_articles_change = f"较昨日 ↑ {rate:.1f}%"
            elif rate < 0:
                hot_articles_change = f"较昨日 ↓ {abs(rate):.1f}%"
            else:
                hot_articles_change = "较昨日不变"

        # 计算hotProjectsChange
        today_projects = HotProjects.objects.filter(type=1, created_time__date=today).count()
        yesterday_projects = HotProjects.objects.filter(type=1, created_time__date=yesterday).count()

        if yesterday_projects == 0:
            hot_projects_change = "较昨日 ↑ 100%" if today_projects > 0 else "较昨日不变"
        else:
            rate = ((today_projects - yesterday_projects) / yesterday_projects) * 100
            if rate > 0:
                hot_projects_change = f"较昨日 ↑ {rate:.1f}%"
            elif rate < 0:
                hot_projects_change = f"较昨日 ↓ {abs(rate):.1f}%"
            else:
                hot_projects_change = "较昨日不变"

        # 计算pushedChange
        total_articles = HotProjects.objects.count()
        if total_articles == 0:
            pushed_change = "已发布占比0%"
        else:
            pushed_ratio = (pushed_total / total_articles) * 100
            pushed_change = f"已发布占比{pushed_ratio:.0f}%"

        # 计算toPushChange
        today_new = HotProjects.objects.filter(created_time__date=today).count()
        if today_new == 0:
            to_push_change = "今日已推送率0%"
        else:
            today_pushed = HotProjects.objects.filter(created_time__date=today, if_sent=1).count()
            to_push_ratio = (today_pushed / today_new) * 100
            to_push_change = f"今日已推送率{to_push_ratio:.0f}%"

        # 计算vulnerabilitiesChange
        month_cve = CveSpider.objects.filter(created_time__date__gte=first_day_of_month).count()
        vulnerabilities_change = f"本月新增{month_cve}个"

        # 计算yesterdayVulnsChange
        yesterday_cve = CveSpider.objects.filter(created_time__date=yesterday).count()

        if yesterday_cve == 0:
            yesterday_vulns_change = "较昨日 ↑ 100%" if today_cve > 0 else "较昨日不变"
        else:
            rate = ((today_cve - yesterday_cve) / yesterday_cve) * 100
            if rate > 0:
                yesterday_vulns_change = f"较昨日 ↑ {rate:.1f}%"
            elif rate < 0:
                yesterday_vulns_change = f"较昨日 ↓ {abs(rate):.1f}%"
            else:
                yesterday_vulns_change = "较昨日不变"

        # 格式化数字，添加千位分隔符
        def format_number(num):
            return f"{num:,}"

        # 构建响应数据
        data = {
            'hotArticlesTotal': format_number(hot_articles_total),
            'hotProjectsTotal': format_number(hot_projects_total),
            'pushedTotal': format_number(pushed_total),
            'toPushTotal': format_number(to_push_total),
            'vulnerabilitiesTotal': format_number(vulnerabilities_total),
            'yesterdayVulns': format_number(today_cve),
            'hotArticlesChange': hot_articles_change,
            'hotProjectsChange': hot_projects_change,
            'pushedChange': pushed_change,
            'toPushChange': to_push_change,
            'vulnerabilitiesChange': vulnerabilities_change,
            'yesterdayVulnsChange': yesterday_vulns_change
        }

        return JsonResponse(data)
    except Exception as e:
        print(f"统计API错误: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# 趋势数据API
from datetime import datetime, timedelta


def trend_data_api(request):
    try:
        # 获取当前日期
        today = datetime.now()
        weeks = []
        articles_data = []
        projects_data = []
        vulnerabilities_data = []

        # 生成8个旬期的日期范围和标签
        for i in range(7, -1, -1):
            # 计算当前旬的开始和结束日期
            current_date = today - timedelta(days=i * 10)
            month = current_date.month
            day = current_date.day

            # 确定旬期
            if day <= 10:
                ten_day_period = f'{month}月上旬'
                start_day = 1
                end_day = 10
            elif day <= 20:
                ten_day_period = f'{month}月中旬'
                start_day = 11
                end_day = 20
            else:
                ten_day_period = f'{month}月下旬'
                start_day = 21
                # 获取当月最后一天
                if month in [4, 6, 9, 11]:
                    end_day = 30
                elif month == 2:
                    # 简单判断闰年
                    if current_date.year % 4 == 0 and (current_date.year % 100 != 0 or current_date.year % 400 == 0):
                        end_day = 29
                    else:
                        end_day = 28
                else:
                    end_day = 31

            # 计算当前旬的开始和结束日期
            start_date = datetime(current_date.year, month, start_day, 0, 0, 0)
            end_date = datetime(current_date.year, month, end_day, 23, 59, 59)

            # 添加旬期标签
            weeks.append(ten_day_period)

            # 查询文章数据 (type=2)
            articles_count = HotProjects.objects.filter(
                type=2,
                created_time__gte=start_date,
                created_time__lte=end_date
            ).count()
            articles_data.append(articles_count)

            # 查询项目数据 (type=1)
            projects_count = HotProjects.objects.filter(
                type=1,
                created_time__gte=start_date,
                created_time__lte=end_date
            ).count()
            projects_data.append(projects_count)

            # 查询漏洞数据
            vulnerabilities_count = CveSpider.objects.filter(
                created_time__gte=start_date,
                created_time__lte=end_date
            ).count()
            vulnerabilities_data.append(vulnerabilities_count)

        # 构建响应数据
        data = {
            'weeks': weeks,
            'articles': articles_data,
            'projects': projects_data,
            'vulnerabilities': vulnerabilities_data
        }

        return JsonResponse(data)
    except Exception as e:
        print(f"趋势数据API错误: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)